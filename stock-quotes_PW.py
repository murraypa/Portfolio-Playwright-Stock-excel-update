import openpyxl
from openpyxl import Workbook
from playwright.sync_api import sync_playwright
from playwright.async_api import async_playwright, TimeoutError

def main():
    with sync_playwright() as p:
        #open browser and load URL outside of loop
        browser = p.chromium.launch(headless=False)
        #browser = p.firefox.launch(headless=False)
        page = browser.new_page()
        #open browser to stock quote page
        page.goto("https://quotes.freerealtime.com/quotes/")

        #open excel workbook
        print('open xl file')
        wb = openpyxl.load_workbook('./Book.xlsx')
        #get worksheet names
        print(f"worksheet names: {wb.sheetnames}")
        #select worksheet with data
        sheet = wb["ticker symbols"]
        #set this worksheet to be active
        ws = wb.active
        #print number of data rows found, minus 1 for header
        number_rows = ws.max_row
        print(f"Data rows found:{number_rows-1}")

        #loop through spreadsheet data rows
        for row in range(2,number_rows+1):
            #get ticker from spreadsheet cell
            ticker = sheet['A'+ str(row)].value
            #click the field to enter ticker
            page.click("#edit-symbol")
            #type in the ticker name
            page.keyboard.insert_text(ticker)
            #press ENTER to search
            page.keyboard.press('Enter')
            #get last, change, percent values using xpath to locate
            last = page.text_content('//*[@id="qmQuoteTable"]/div[2]/div/div/div/div[1]/div/div[2]/div/div/div[1]/div[1]/span[1]')
            change = page.text_content('//*[@id="qmQuoteTable"]/div[2]/div/div/div/div[1]/div/div[2]/div/div/div[1]/div[1]/span[2]/span[2]')
            percent = page.text_content('//*[@id="qmQuoteTable"]/div[2]/div/div/div/div[1]/div/div[2]/div/div/div[1]/div[1]/span[2]/span[4]')
            print(f"ticker={ticker}\t last={last}\t change={change}\t percent={percent}")
            #input("press enter to continue")

            #update 3 cells in this row of the spreadsheet
            ws.cell(row,2,last)
            ws.cell(row,3,change)
            ws.cell(row,4,percent)

        #save spreadsheet
        wb.save('./Book.xlsx')
        #close the browser window
        browser.close()



if __name__ == '__main__':
    main()
