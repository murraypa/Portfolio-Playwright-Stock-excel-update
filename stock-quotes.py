from multiprocessing.connection import wait
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.service import Service

import openpyxl
from openpyxl import Workbook

print('open xl file')
#open excel workbook
wb = openpyxl.load_workbook('./Book.xlsx')
#get worksheet names
print(f"worksheet names: {wb.sheetnames}")
#get single worksheet
sheet = wb["ticker symbols"]
print(f"sheet={sheet}")
#set active worksheet
ws = wb.active
ticker = sheet['A2'].value
print(f"row1={ticker}\n")
number_rows = ws.max_row
print(f"max rows in sheet:{number_rows}")

#control web browser with web driver
#driver = webdriver.Chrome(".\chromedriver.exe")
#driver = webdriver.chrome()

driver = webdriver.Firefox(executable_path=r'C:\\Users\user\\Downloads\\python\\Selenium\\selenium-tests\\geckodriver.exe')
driver.get("https://www.nasdaq.com/market-activity/quotes/real-time")
#driver = Service(r'C:\\Users\\user\\Downloads\\python\\Selenium\\selenium-tests\\geckodriver.exe')
#driver = webdriver.Firefox(executable_path=r'C:\Users\user\Downloads\python\Selenium\selenium-tests\geckodriver.exe')
print("sleeping 10")
sleep(10)
quotelookup = driver.find_element(By.CSS_SELECTOR,'#find-symbol-input-dark').send_keys(ticker)
quotelookup = driver.find_element(By.CSS_SELECTOR,'#find-symbol-input-dark').send_keys(Keys.RETURN)
print("sleeping 10")
sleep(10)
NLSvolume = driver.find_element(By.CSS_SELECTOR,'.real-time-trades-info__cell--value')
print(f"NLS volume={NLSvolume}")
# quotelookup = driver.find_element(By.XPATH,'//*[@id="yfin-usr-qry"]').send_keys(ticker)
# quotelookup = driver.find_element(By.XPATH,'//*[@id="yfin-usr-qry"]').send_keys(Keys.RETURN)
# quotelookup = driver.find_element(By.XPATH,'//*[@id="yfin-usr-qry"]').send_keys(Keys.RETURN)


#loop through all rows, skipping title row 1
for symbol in range(2,number_rows+1):
    ticker = sheet['A'+ str(symbol)].value
    print(f"ticker={ticker}")
    #put current ticker into "quote lookup" box
    # quotelookup = driver.find_element(By.CSS_SELECTOR,'input.D\(ib\)').send_keys(ticker)
    # quotelookup = driver.find_element(By.CSS_SELECTOR,'input.D\(ib\)').send_keys(Keys.ENTER)

